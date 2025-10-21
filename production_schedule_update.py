import openpyxl
import gspread
import datetime
import os
import sys
import traceback
from dotenv import load_dotenv
import smtplib
from email.mime.text import MIMEText
import warnings
import time
import tempfile
import shutil
import zipfile
import xml.etree.ElementTree as ET
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# openpyxlの日付警告を抑制
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# .envファイルから環境変数を読み込み
load_dotenv()

# --- メール通知用の設定 ---
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVERS = os.getenv("EMAIL_RECEIVERS", "").split(",") if os.getenv("EMAIL_RECEIVERS") else []
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.office365.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))

# Google API設定
GOOGLE_SERVICE_ACCOUNT_KEY_FILE = os.getenv("GOOGLE_SERVICE_ACCOUNT_KEY_FILE")
GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID", "")  # アップロード先フォルダID（オプション）

# Google API設定の存在確認
if not GOOGLE_SERVICE_ACCOUNT_KEY_FILE:
    raise ValueError("GOOGLE_SERVICE_ACCOUNT_KEY_FILE が .env ファイルに設定されていません")

# Google API認証ファイルパスの解決（環境非依存）
def resolve_google_api_key_file(filename):
    """
    Google API認証ファイルのパスを環境非依存で解決する
    """
    # 1. 現在の作業ディレクトリからの相対パス
    if os.path.exists(filename):
        return os.path.abspath(filename)
    
    # 2. スクリプトと同じディレクトリ
    script_dir = os.path.dirname(os.path.abspath(__file__))
    script_relative_path = os.path.join(script_dir, filename)
    if os.path.exists(script_relative_path):
        return os.path.abspath(script_relative_path)
    
    # 3. ファイルが見つからない場合
    raise FileNotFoundError(
        f"Google API認証ファイルが見つかりません: {filename}\n"
        f"検索パス:\n"
        f"  - 現在の作業ディレクトリ: {os.path.abspath(filename)}\n"
        f"  - スクリプトディレクトリ: {script_relative_path}"
    )

# Google API認証ファイルパスを解決
RESOLVED_GOOGLE_API_KEY_FILE = resolve_google_api_key_file(GOOGLE_SERVICE_ACCOUNT_KEY_FILE)

def send_error_email(error_info):
    """
    エラー発生時に指定されたアカウントへメールを送信する関数
    """
    try:
        # プログラム名とファイルパスを取得
        program_name = os.path.basename(sys.argv[0])
        file_path = os.path.abspath(sys.argv[0])
        
        # 件名にプログラム名を追記
        subject = f"【エラー通知】{program_name} 実行中にエラーが発生しました"
        
        # 本文にプログラム名とファイルパスを追記
        body = f"""
お疲れ様です。

Pythonスクリプトの実行中にエラーが発生しました。
下記に詳細を記載します。

---
プログラム名: {program_name}

ファイルパス: {file_path}

日時: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')}

エラー詳細:
{error_info}
---

お手数ですが、ご確認をお願いします。
"""
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = EMAIL_SENDER
        msg["To"] = ", ".join(EMAIL_RECEIVERS)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECEIVERS, msg.as_string())
        print("エラー通知メールを送信しました。")

    except Exception as e:
        print(f"メール送信中にエラーが発生しました: {e}")

def detect_hidden_columns(xlsx_path, sheet_name):
    """openpyxl と XML を組み合わせた非表示列検出"""
    hidden_columns = []
    
    try:
        # openpyxl による検出
        wb_format = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=False, keep_links=False)
        source_ws_format = wb_format[sheet_name]
        
        for col_letter, col_dim in source_ws_format.column_dimensions.items():
            if col_dim.hidden:
                hidden_columns.append(col_letter)
        
        wb_format.close()
        
        # XML による検出
        with zipfile.ZipFile(xlsx_path, 'r') as zip_file:
            workbook_xml = zip_file.read('xl/workbook.xml')
            workbook_root = ET.fromstring(workbook_xml)
            
            sheet_mapping = {}
            for sheet in workbook_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                sheet_name_xml = sheet.get('name')
                r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                sheet_mapping[sheet_name_xml] = r_id
            
            rels_xml = zip_file.read('xl/_rels/workbook.xml.rels')
            rels_root = ET.fromstring(rels_xml)
            
            sheet_files = {}
            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                rel_id = rel.get('Id')
                target = rel.get('Target')
                if 'worksheets' in target:
                    sheet_files[rel_id] = target
            
            if sheet_name in sheet_mapping:
                r_id = sheet_mapping[sheet_name]
                if r_id in sheet_files:
                    sheet_file = sheet_files[r_id]
                    try:
                        sheet_xml = zip_file.read(f'xl/{sheet_file}')
                        sheet_root = ET.fromstring(sheet_xml)
                        
                        cols = sheet_root.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}cols')
                        if cols is not None:
                            for col in cols.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}col'):
                                min_col = int(col.get('min', 1))
                                max_col = int(col.get('max', min_col))
                                hidden = col.get('hidden')
                                
                                for col_num in range(min_col, max_col + 1):
                                    col_letter = openpyxl.utils.get_column_letter(col_num)
                                    if hidden == "1" and col_letter not in hidden_columns:
                                        hidden_columns.append(col_letter)
                    except Exception:
                        pass
    
    except Exception:
        pass
    
    return hidden_columns

def extract_excel_sheet_to_temp_file(excel_file_path, sheet_name):
    """
    Excelファイルから指定シートを抽出して一時ファイルを作成する関数
    読み取り専用で元ファイルを変更しない
    """
    try:
        # データ取得用：読み取り専用で開く
        wb_data = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True, keep_links=False)
        
        # 書式取得用：通常モードで開く（書式情報のため）
        wb_format = openpyxl.load_workbook(excel_file_path, read_only=False, data_only=False, keep_links=False)
        
        # 指定シートが存在するかチェック
        if sheet_name not in wb_data.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が見つかりません。利用可能なシート: {wb_data.sheetnames}")
        
        # 新しいワークブックを作成
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name
        
        # データ用シートと書式用シートを取得
        source_ws_data = wb_data[sheet_name]
        source_ws_format = wb_format[sheet_name]
        
        # 非表示列検出（openpyxl + XML）
        hidden_columns = detect_hidden_columns(excel_file_path, sheet_name)
        
        # データをコピー（読み取り専用から）
        data_rows = list(source_ws_data.iter_rows(values_only=True))
        
        # 書式情報をコピー
        for row_idx, row in enumerate(source_ws_format.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if row_idx <= len(data_rows) and col_idx <= len(data_rows[row_idx-1]):
                    # セルに値を設定
                    new_cell = new_ws.cell(row=row_idx, column=col_idx, value=data_rows[row_idx-1][col_idx-1])
                    
                    # 書式をコピー（StyleProxyエラーを回避）
                    try:
                        if cell.has_style:
                            # フォント
                            if cell.font:
                                new_cell.font = openpyxl.styles.Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    vertAlign=cell.font.vertAlign,
                                    underline=cell.font.underline,
                                    strike=cell.font.strike,
                                    color=cell.font.color
                                )
                            
                            # 境界線
                            if cell.border:
                                new_cell.border = openpyxl.styles.Border(
                                    left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom
                                )
                            
                            # 塗りつぶし
                            if cell.fill:
                                new_cell.fill = openpyxl.styles.PatternFill(
                                    fill_type=cell.fill.fill_type,
                                    start_color=cell.fill.start_color,
                                    end_color=cell.fill.end_color
                                )
                            
                            # 数値フォーマット
                            if cell.number_format:
                                new_cell.number_format = cell.number_format
                            
                            # 配置
                            if cell.alignment:
                                new_cell.alignment = openpyxl.styles.Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    text_rotation=cell.alignment.text_rotation,
                                    wrap_text=cell.alignment.wrap_text,
                                    shrink_to_fit=cell.alignment.shrink_to_fit,
                                    indent=cell.alignment.indent
                                )
                    except Exception as style_error:
                        # 書式コピーに失敗した場合はスキップ
                        continue
        
        # 列幅をコピー
        for col_letter in source_ws_format.column_dimensions:
            new_ws.column_dimensions[col_letter].width = source_ws_format.column_dimensions[col_letter].width
            new_ws.column_dimensions[col_letter].hidden = source_ws_format.column_dimensions[col_letter].hidden
        
        # 行の高さをコピー
        for row_num in source_ws_format.row_dimensions:
            new_ws.row_dimensions[row_num].height = source_ws_format.row_dimensions[row_num].height
            new_ws.row_dimensions[row_num].hidden = source_ws_format.row_dimensions[row_num].hidden
        
        # 一時ファイルを作成
        temp_dir = tempfile.mkdtemp()
        temp_file_path = os.path.join(temp_dir, f"{sheet_name}.xlsx")
        
        # 一時ファイルに保存
        new_wb.save(temp_file_path)
        new_wb.close()
        wb_data.close()
        wb_format.close()
        
        return temp_file_path, temp_dir
        
    except Exception as e:
        raise Exception(f"Excelシート抽出中にエラーが発生しました: {e}")

def get_google_drive_service():
    """
    Google Drive APIサービスを取得する関数
    """
    try:
        # 認証スコープ
        SCOPES = [
            'https://www.googleapis.com/auth/drive',
            'https://www.googleapis.com/auth/spreadsheets'
        ]
        
        # 認証情報をロード
        credentials = service_account.Credentials.from_service_account_file(
            RESOLVED_GOOGLE_API_KEY_FILE, scopes=SCOPES
        )
        
        # Google Drive APIサービスを構築
        service = build('drive', 'v3', credentials=credentials)
        return service
        
    except Exception as e:
        raise Exception(f"Google Drive API認証中にエラーが発生しました: {e}")

def find_existing_file(service, filename, target_spreadsheet_id=None):
    """
    既存のファイルを検索する関数
    """
    try:
        # 特定のスプレッドシートIDが指定されている場合はそれを使用
        if target_spreadsheet_id:
            # ファイルが存在するかチェック
            try:
                file_info = service.files().get(fileId=target_spreadsheet_id, fields='id,name').execute()
                print(f"指定されたスプレッドシートが見つかりました: {file_info.get('name')}")
                return target_spreadsheet_id
            except Exception as e:
                print(f"指定されたスプレッドシートIDが見つかりません: {target_spreadsheet_id}")
                return None
        
        # 通常の検索処理
        query = f"name='{filename}' and trashed=false"
        if GOOGLE_DRIVE_FOLDER_ID:
            query += f" and parents in '{GOOGLE_DRIVE_FOLDER_ID}'"
        
        results = service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get('files', [])
        
        if files:
            return files[0]['id']  # 最初に見つかったファイルのIDを返す
        return None
        
    except Exception as e:
        print(f"既存ファイル検索中にエラーが発生しました: {e}")
        return None

def upload_to_google_drive(service, temp_file_path, filename, target_spreadsheet_id=None):
    """
    Googleドライブにファイルをアップロードする関数（上書き対応）
    """
    try:
        # 既存ファイルを検索
        existing_file_id = find_existing_file(service, filename, target_spreadsheet_id)
        
        # ファイルメタデータ
        file_metadata = {
            'name': filename
        }
        
        # フォルダIDが指定されている場合は親フォルダを設定
        if GOOGLE_DRIVE_FOLDER_ID:
            file_metadata['parents'] = [GOOGLE_DRIVE_FOLDER_ID]
        
        # メディアファイル
        media = MediaFileUpload(temp_file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        if existing_file_id and existing_file_id != target_spreadsheet_id:
            # 既存のExcelファイルを更新（スプレッドシートとは別のファイル）
            update_metadata = {'name': filename}
            file = service.files().update(
                fileId=existing_file_id,
                body=update_metadata,
                media_body=media
            ).execute()
            return file.get('id')
        else:
            # 新規ファイルを作成
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            return file.get('id')
        
    except Exception as e:
        raise Exception(f"Googleドライブアップロード中にエラーが発生しました: {e}")

def convert_to_google_sheets(service, file_id, target_spreadsheet_id=None, temp_file_path=None):
    """
    ExcelファイルをGoogleスプレッドシートに変換する関数
    """
    try:
        if target_spreadsheet_id:
            # 既存のスプレッドシートを更新
            # 既存のスプレッドシートの内容をクリア
            try:
                gspread_client = gspread.service_account(filename=RESOLVED_GOOGLE_API_KEY_FILE)
                spreadsheet = gspread_client.open_by_key(target_spreadsheet_id)
                worksheet = spreadsheet.sheet1
                worksheet.clear()
            except Exception as e:
                print(f"スプレッドシートクリア中にエラー: {e}")
            
            # 新しい内容をアップロード
            media = MediaFileUpload(temp_file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            service.files().update(
                fileId=target_spreadsheet_id,
                media_body=media
            ).execute()
            
            # 一時Excelファイルを削除（確実に削除）
            if file_id != target_spreadsheet_id:  # スプレッドシートとは別のファイルの場合のみ削除
                try:
                    service.files().delete(fileId=file_id).execute()
                    print("一時Excelファイルを削除しました")
                except Exception as delete_error:
                    print(f"一時Excelファイル削除をスキップしました: {delete_error}")
            
            spreadsheet_id = target_spreadsheet_id
        else:
            # 新しいスプレッドシートを作成
            converted_file = service.files().copy(
                fileId=file_id,
                body={
                    'name': '量産品管理日程_不二工機',
                    'mimeType': 'application/vnd.google-apps.spreadsheet'
                }
            ).execute()
            
            spreadsheet_id = converted_file.get('id')
            
            # 元のExcelファイルを削除
            service.files().delete(fileId=file_id).execute()
        
        # 列の表示/非表示と列幅を設定
        try:
            gspread_client = gspread.service_account(filename=RESOLVED_GOOGLE_API_KEY_FILE)
            spreadsheet = gspread_client.open_by_key(spreadsheet_id)
            worksheet = spreadsheet.sheet1
            sheet_id = worksheet.id
            
            hidden_columns = detect_hidden_columns(EXCEL_FILE_PATH, EXCEL_SHEET_NAME)
            
            all_values = worksheet.get_all_values()
            if all_values:
                max_cols = len(all_values[0]) if all_values[0] else 0
                
                for col_num in range(1, max_cols + 1):
                    col_letter = openpyxl.utils.get_column_letter(col_num)
                    pixel_size = 70 if col_num >= 3 else 100
                    is_hidden = col_letter in hidden_columns
                    
                    try:
                        format_request = {
                            "requests": [{
                                "updateDimensionProperties": {
                                    "range": {
                                        "sheetId": sheet_id,
                                        "dimension": "COLUMNS",
                                        "startIndex": col_num - 1,
                                        "endIndex": col_num
                                    },
                                    "properties": {
                                        "pixelSize": pixel_size,
                                        "hiddenByUser": is_hidden
                                    },
                                    "fields": "pixelSize,hiddenByUser"
                                }
                            }]
                        }
                        spreadsheet.batch_update(format_request)
                    except Exception:
                        pass
        except Exception:
            pass
        
        # GoogleスプレッドシートのURLを表示
        spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
        print(f"GoogleスプレッドシートURL: {spreadsheet_url}")
        
        return spreadsheet_id
        
    except Exception as e:
        raise Exception(f"Googleスプレッドシート変換中にエラーが発生しました: {e}")


def cleanup_temp_files(temp_dir):
    """
    一時ファイルとディレクトリを削除する関数
    """
    try:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
    except Exception as e:
        print(f"一時ファイル削除中にエラーが発生しました: {e}")

# --- メイン処理 ---
try:
    # 設定値
    EXCEL_FILE_PATH = r"\\192.168.1.200\共有\生産管理課\量産品管理日程.xlsx"
    EXCEL_SHEET_NAME = "不二工機"
    GOOGLE_FILE_NAME = "量産品管理日程_不二工機"
    
    # 新しいフォルダID（URLから抽出）
    GOOGLE_DRIVE_FOLDER_ID = "15XI2hV6hwsxsXvo1Y4oihBHVsSrkPpv0"
    
    # 指定されたスプレッドシートID（上書き対象）
    TARGET_SPREADSHEET_ID = "11PvactpesgyOgMFlufe8FcZ3JShTOgEBAmsUaGk3XLE"
    
    print("量産品日程更新処理を開始します...")
    
    # 1. Excelファイルから対象シートを抽出して一時ファイルを作成
    temp_file_path, temp_dir = extract_excel_sheet_to_temp_file(EXCEL_FILE_PATH, EXCEL_SHEET_NAME)
    
    # 2. Google Drive APIサービスを取得
    drive_service = get_google_drive_service()
    
    # 3. Googleドライブにアップロード
    file_id = upload_to_google_drive(drive_service, temp_file_path, GOOGLE_FILE_NAME, TARGET_SPREADSHEET_ID)
    
    # 4. Googleスプレッドシートに変換
    spreadsheet_id = convert_to_google_sheets(drive_service, file_id, TARGET_SPREADSHEET_ID, temp_file_path)
    
    # 5. 一時ファイルをクリーンアップ
    cleanup_temp_files(temp_dir)
    
    print(f"量産品日程更新処理が完了しました。")
    
except Exception as e:
    error_detail = traceback.format_exc()
    print(f"エラーが発生しました: {e}")
    print(f"詳細: {error_detail}")
    
    # 一時ファイルのクリーンアップ
    try:
        if 'temp_dir' in locals():
            cleanup_temp_files(temp_dir)
    except:
        pass
    
    # エラー通知メールを送信
    if EMAIL_SENDER and EMAIL_PASSWORD and EMAIL_RECEIVERS:
        send_error_email(error_detail)
    else:
        print("メール通知設定が不完全なため、エラー通知メールは送信されませんでした。")
    
    # エラーを再発生させてプログラムを停止
    raise