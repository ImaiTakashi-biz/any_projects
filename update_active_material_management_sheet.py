import openpyxl
import gspread
import datetime
from oauth2client.service_account import ServiceAccountCredentials
import smtplib
from email.mime.text import MIMEText
import traceback
import os
import sys
from dotenv import load_dotenv

# .envファイルから環境変数を読み込み
load_dotenv()

# 通知アカウント
# ・takada@araiseimitsu.onmicrosoft.com
# ・imai@araiseimitsu.onmicrosoft.com
# ・n.kizaki@araiseimitsu.onmicrosoft.com

# --- メール通知用の設定 ---
# これらの設定値は、.envファイルから読み込まれます。
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVERS = os.getenv("EMAIL_RECEIVERS", "").split(",") if os.getenv("EMAIL_RECEIVERS") else []
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.office365.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))

# Google API設定
GOOGLE_SERVICE_ACCOUNT_KEY_FILE = os.getenv("GOOGLE_SERVICE_ACCOUNT_KEY_FILE")

# Google API設定の存在確認
if not GOOGLE_SERVICE_ACCOUNT_KEY_FILE:
    raise ValueError("GOOGLE_SERVICE_ACCOUNT_KEY_FILE が .env ファイルに設定されていません")

# Google API認証ファイルパスの解決（環境非依存）
def resolve_google_api_key_file(filename):
    """
    Google API認証ファイルのパスを環境非依存で解決する
    """
    # 1. 現在の作業ディレクトリからの相対パス
    if os.path.exists(filename):
        return os.path.abspath(filename)
    
    # 2. スクリプトと同じディレクトリ
    script_dir = os.path.dirname(os.path.abspath(__file__))
    script_relative_path = os.path.join(script_dir, filename)
    if os.path.exists(script_relative_path):
        return script_relative_path
    
    # 3. ファイルが見つからない場合
    raise FileNotFoundError(
        f"Google API認証ファイルが見つかりません: {filename}\n"
        f"検索パス:\n"
        f"  - 現在の作業ディレクトリ: {os.path.abspath(filename)}\n"
        f"  - スクリプトディレクトリ: {script_relative_path}"
    )

# Google API認証ファイルパスを解決
RESOLVED_GOOGLE_API_KEY_FILE = resolve_google_api_key_file(GOOGLE_SERVICE_ACCOUNT_KEY_FILE)

# メール設定の存在確認
if not EMAIL_SENDER:
    raise ValueError("EMAIL_SENDER が .env ファイルに設定されていません")
if not EMAIL_PASSWORD:
    raise ValueError("EMAIL_PASSWORD が .env ファイルに設定されていません")
if not EMAIL_RECEIVERS:
    raise ValueError("EMAIL_RECEIVERS が .env ファイルに設定されていません")

def send_error_email(error_info):
    """
    エラー発生時に指定されたアカウントへメールを送信する関数
    """
    try:
        # プログラム名とファイルパスを取得
        program_name = os.path.basename(sys.argv[0])
        file_path = os.path.abspath(sys.argv[0])
        
        # 件名にプログラム名を追記
        subject = f"【エラー通知】{program_name} 実行中にエラーが発生しました"
        
        # 本文にプログラム名とファイルパスを追記
        body = f"""
お疲れ様です。

Pythonスクリプトの実行中にエラーが発生しました。
下記に詳細を記載します。

---
プログラム名: {program_name}

ファイルパス: {file_path}

日時: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')}

エラー詳細:
{error_info}
---

お手数ですが、ご確認をお願いします。
"""
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = EMAIL_SENDER
        msg["To"] = ", ".join(EMAIL_RECEIVERS)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECEIVERS, msg.as_string())
        print("エラー通知メールを送信しました。")

    except Exception as e:
        print(f"メール送信中にエラーが発生しました: {e}")

# --- メイン処理 ---
try:
    # 認証情報を設定
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/spreadsheets",
             "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(RESOLVED_GOOGLE_API_KEY_FILE, scope)
    client = gspread.authorize(creds)

    # スプレッドシートを開く
    ws = client.open_by_key("1pJluvzitGCr4i_kIvUThczh_6YcqKg6ujZplTz_JR1g") # スプレッドシートのkey
    sh = ws.worksheet("Data")
    sh1 = ws.worksheet("稼働中材料管理表")

    file_pass = r"\\192.168.1.200\共有\生産管理課\セット予定表.xlsx"
    wb1 = openpyxl.load_workbook(file_pass, read_only=True)
    ws1 = wb1["生産中"]
    last_row = ws1.max_row

    for _ in range(0 + 1, last_row + 1):
        if ws1.cell(_, 4).value is not None:
            last_row1 = _

    two_list = []
    for _ in range(0 + 1, last_row1 + 1):
        row_list4 = []
        value1 = ws1.cell(_, 4).value
        if isinstance(value1, datetime.datetime):
            value1 = value1.strftime("%Y/%m/%d")
        row_list4.append(value1)
        row_list4.append(ws1.cell(_, 5).value)
        row_list4.append(ws1.cell(_, 9).value)
        row_list4.append(ws1.cell(_, 11).value)
        row_list4.append(ws1.cell(_, 12).value)
        row_list4.append(ws1.cell(_, 24).value)
        if _ != 1 and isinstance(ws1.cell(_, 27).value, float):
            row_list4.append(round(float(ws1.cell(_, 27).value)))
        two_list.append(row_list4)

    # すべての要素を文字列に変換
    def convert_to_serializable(obj):
        if isinstance(obj, datetime.datetime):
            return obj.strftime("%Y/%m/%d")
        elif obj is None:
            return ""
        return obj

    two_list_serializable = [[convert_to_serializable(item) for item in sublist] for sublist in two_list]

    sh.clear()
    ws.values_append("Data", {"valueInputOption": "USER_ENTERED"}, {"values": two_list_serializable})
    wb1.close()

    print("完了しました。")

except Exception as e:
    error_detail = traceback.format_exc()
    send_error_email(error_detail)
    # エラーを再発生させてプログラムを停止
    raise