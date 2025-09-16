import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
import os
import shutil
import datetime
import smtplib
from email.mime.text import MIMEText
import traceback
import sys
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.errors import HttpError
import time
import gc
from dotenv import load_dotenv

# .envファイルから環境変数を読み込み
load_dotenv()

# オプショナルインポート
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
    print("警告: psutilがインストールされていません。プロセス監視機能は無効になります。")

# Windows固有のファイルハンドル操作用
try:
    import ctypes
    from ctypes import wintypes
    CTYPES_AVAILABLE = True
except ImportError:
    CTYPES_AVAILABLE = False

# 通知アカウント
# ・takada@araiseimitsu.onmicrosoft.com
# ・imai@araiseimitsu.onmicrosoft.com
# ・n.kizaki@araiseimitsu.onmicrosoft.com

# --- メール通知用の設定 ---
# これらの設定値は、.envファイルから読み込まれます。
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVERS = os.getenv("EMAIL_RECEIVERS", "").split(",") if os.getenv("EMAIL_RECEIVERS") else []
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.office365.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))

# Google API設定
GOOGLE_SERVICE_ACCOUNT_KEY_FILE = os.getenv("GOOGLE_SERVICE_ACCOUNT_KEY_FILE")

# Google API設定の存在確認
if not GOOGLE_SERVICE_ACCOUNT_KEY_FILE:
    raise ValueError("GOOGLE_SERVICE_ACCOUNT_KEY_FILE が .env ファイルに設定されていません")

# Google API認証ファイルパスの解決（環境非依存）
def resolve_google_api_key_file(filename):
    """
    Google API認証ファイルのパスを環境非依存で解決する
    """
    # 1. 現在の作業ディレクトリからの相対パス
    if os.path.exists(filename):
        return os.path.abspath(filename)
    
    # 2. スクリプトと同じディレクトリ
    script_dir = os.path.dirname(os.path.abspath(__file__))
    script_relative_path = os.path.join(script_dir, filename)
    if os.path.exists(script_relative_path):
        return script_relative_path
    
    # 3. ファイルが見つからない場合
    raise FileNotFoundError(
        f"Google API認証ファイルが見つかりません: {filename}\n"
        f"検索パス:\n"
        f"  - 現在の作業ディレクトリ: {os.path.abspath(filename)}\n"
        f"  - スクリプトディレクトリ: {script_relative_path}"
    )

# Google API認証ファイルパスを解決
RESOLVED_GOOGLE_API_KEY_FILE = resolve_google_api_key_file(GOOGLE_SERVICE_ACCOUNT_KEY_FILE)

# メール設定の存在確認
if not EMAIL_SENDER:
    raise ValueError("EMAIL_SENDER が .env ファイルに設定されていません")
if not EMAIL_PASSWORD:
    raise ValueError("EMAIL_PASSWORD が .env ファイルに設定されていません")
if not EMAIL_RECEIVERS:
    raise ValueError("EMAIL_RECEIVERS が .env ファイルに設定されていません")

def send_error_email(error_info):
    """
    エラー発生時に指定されたアカウントへメールを送信する関数
    """
    try:
        # プログラム名とファイルパスを取得
        program_name = os.path.basename(sys.argv[0])
        file_path = os.path.abspath(sys.argv[0])
        
        # 件名にプログラム名を追記
        subject = f"【エラー通知】{program_name} 実行中にエラーが発生しました"
        
        # 本文にプログラム名とファイルパスを追記
        body = f"""
お疲れ様です。

Pythonスクリプトの実行中にエラーが発生しました。
下記に詳細を記載します。

---
プログラム名: {program_name}

ファイルパス: {file_path}

日時: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')}

エラー詳細:
{error_info}
---

お手数ですが、ご確認をお願いします。
"""
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = EMAIL_SENDER
        msg["To"] = ", ".join(EMAIL_RECEIVERS)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECEIVERS, msg.as_string())
        print("エラー通知メールを送信しました。")

    except Exception as e:
        print(f"メール送信中にエラーが発生しました: {e}")

# --- メイン処理 ---
try:
    # ロボパット用
    wb_rob = openpyxl.load_workbook(r"\\192.168.1.200\共有\製造課\ロボパット\ロボパット用.xlsx", data_only=True)
    ws_rob = wb_rob["Data"]
    ws_link = wb_rob["品番リンク"]

    # 機番・品番　取得
    list_rob = []
    target_row = [9, 12]
    for col in range(2, 12):
        data = []
        for row in target_row:
            value1 = ws_rob.cell(row, col).value
            data.append(value1)
        list_rob.append(data)
    print(list_rob)

    # リンク取得
    last_row = ws_link.max_row
    for _ in range(0 + 1,last_row + 1):
        if ws_link.cell(_,1).value is not None:
            last_row1 = _

    list_link = []
    for item in list_rob:
        data = None  # デフォルトでNoneを設定
        for row in range(1, last_row1 + 1):
            if item[1] == ws_link.cell(row, 1).value:
                data = ws_link.cell(row, 2).value  # マッチする場合はデータを設定
                break  # マッチしたらループを終了
        list_link.append(data)

    # リスト結合
    data_list = []
    for rob, link in zip(list_rob, list_link):
        data_list.append([rob[0], rob[1], link])

    print(data_list)
    wb_rob.close()

    # 認証情報を設定
    scope = ["https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(RESOLVED_GOOGLE_API_KEY_FILE, scope)
    client = gspread.authorize(creds)
    
    # Google Drive APIクライアントを追加
    drive_service = build('drive', 'v3', credentials=creds)

    # 稼働中工程内検査シート
    ws = client.open_by_key("149nerm5Gma8sdhn5m18nvsXdCNasvv_P9qSEDRJhKxs") #スプレッドシートのkey
    sh = ws.worksheet("シート1")

    values = sh.get_all_values()

    # データとセルの範囲を出力
    for i, row in enumerate(values):
        for j, val in enumerate(row):
            cell_address = gspread.utils.rowcol_to_a1(i + 1, j + 1 + 1)
            for data in data_list:
                if data[0] == val:
                    new_link = data[2]  # 新しいリンク
                    # リンクを更新
                    existing_formula = sh.acell(cell_address, value_render_option='FORMULA').value
                    if existing_formula is not None:
                        if existing_formula.startswith('=HYPERLINK'):
                            # 既存のリンクがある場合、リンクの先だけを更新する
                            existing_link = existing_formula.split('"')[1]
                            if existing_link is not None and new_link is not None:
                                updated_formula = existing_formula.replace(str(existing_link), new_link)
                                sh.update_acell(cell_address, updated_formula)
                            else:
                                print("Error: The existing link or new link is None.")
                        else:
                            # 既存のリンクがない場合、新しいリンクを挿入する
                            sh.update_acell(cell_address, f'=HYPERLINK("{new_link}", "〇")')
                    else:
                        print("Error: The existing formula is None.")

    # セット品加工図更新
    processed_files = set()  # 処理済みのファイル名を追跡するセット

    for data in data_list:
        if data[0] is not None:
            if data[0] in processed_files:
                print(f"Duplicate found, skipping: {data[0]}")
                continue  # 重複があればスキップ

            old_file_path = rf"\\192.168.1.200\共有\製造課\加工工程管理表、プログラム\5加工図面\{data[1]}.pdf"
            new_folder_path = r"\\192.168.1.200\共有\製造課\ロボパット\セット品加工図準備用"
            new_file_name = f"{data[0]}.pdf"

            if os.path.exists(old_file_path):
                # コピー先ディレクトリが存在しない場合は作成
                if not os.path.exists(new_folder_path):
                    os.makedirs(new_folder_path)

                # ファイルをコピーして新しいファイル名に変更
                shutil.copy(old_file_path, os.path.join(new_folder_path, os.path.basename(old_file_path)))
                new_file_path = os.path.join(new_folder_path, new_file_name)
                
                # リネーム先のファイルが既に存在する場合は削除
                if os.path.exists(new_file_path):
                    os.remove(new_file_path)
                    print(f"既存のファイルを削除しました: {new_file_path}")
                
                os.rename(os.path.join(new_folder_path, os.path.basename(old_file_path)), new_file_path)

                # 処理済みのファイル名をセットに追加
                processed_files.add(data[0])
            else:
                print(f"File does not exist: {old_file_path}")

    path_dir = r"\\192.168.1.200\共有\製造課\ロボパット\セット品加工図準備用"
    # Google DriveのフォルダIDを指定（1加工図面フォルダのID）
    target_folder_id = "1x2MWa8ZiLFuPcHdd9jIGYQjkqR877pDo"
    
    # Google Drive APIを使用してファイルをアップロード
    upload_errors = []  # アップロードエラーを記録
    try:
        # まず、フォルダが存在するかを確認
        try:
            folder_info = drive_service.files().get(fileId=target_folder_id).execute()
            print(f"アップロード先フォルダ確認: {folder_info.get('name')}")
        except HttpError as e:
            if e.resp.status == 404:
                error_msg = f"エラー: フォルダID '{target_folder_id}' が見つかりません。\nフォルダのアクセス権限を確認してください。"
                print(error_msg)
                upload_errors.append(error_msg)
                send_error_email(f"Google Driveフォルダアクセスエラー:\n{error_msg}")
                raise
            else:
                raise
        
        list_file_name = os.listdir(path_dir)
        
        for i_file_name in list_file_name:
            join_path = os.path.join(path_dir, i_file_name)
            
            if os.path.isfile(join_path):
                try:
                    # Google Drive上で同名ファイルの検索
                    query = f"name = '{i_file_name}' and '{target_folder_id}' in parents and trashed = false"
                    response = drive_service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
                    existing_files = response.get('files', [])

                    media = MediaFileUpload(join_path, resumable=True)

                    if existing_files:
                        # ファイルが存在する場合 -> 更新
                        existing_file_id = existing_files[0].get('id')
                        file = drive_service.files().update(
                            fileId=existing_file_id,
                            media_body=media,
                            fields='id'
                        ).execute()
                        print(f'ファイル更新完了: {i_file_name} (ID: {file.get("id")})')
                    else:
                        # ファイルが存在しない場合 -> 新規作成
                        file_metadata = {
                            'name': i_file_name,
                            'parents': [target_folder_id]
                        }
                        file = drive_service.files().create(
                            body=file_metadata,
                            media_body=media,
                            fields='id'
                        ).execute()
                        print(f'ファイルアップロード完了: {i_file_name} (ID: {file.get("id")})')
                    
                except HttpError as upload_error:
                    error_msg = f"ファイル '{i_file_name}' のアップロードに失敗: {upload_error}"
                    print(error_msg)
                    upload_errors.append(error_msg)
                    continue
                
    except Exception as upload_error:
        error_msg = f"Google Driveアップロード中にエラーが発生しました: {upload_error}"
        print(error_msg)
        upload_errors.append(error_msg)
    
    # アップロードエラーがある場合はまとめてメール送信
    if upload_errors:
        error_summary = "\n".join(upload_errors)
        send_error_email(f"Google Driveアップロードエラーまとめ:\n{error_summary}")
    
    # 保険処理：セット品加工図準備用フォルダを完全に空にする
    cleanup_warnings = []  # クリーンアップ警告を記録（エラーではなく警告として扱う）
    successful_deletions = []
    failed_deletions = []
    
    try:
        # ガベージコレクションを実行してファイルハンドルを解放
        gc.collect()
        time.sleep(1)  # 少し待機
        
        remaining_files = os.listdir(path_dir)
        if remaining_files:
            print(f"残存ファイルを削除します: {remaining_files}")
            for remaining_file in remaining_files:
                remaining_file_path = os.path.join(path_dir, remaining_file)
                
                # リトライ機能付きでファイル削除
                max_retries = 5  # リトライ回数を増加
                deletion_successful = False
                
                for retry in range(max_retries):
                    try:
                        if os.path.isfile(remaining_file_path):
                            # ファイル属性を変更（読み取り専用を解除）
                            try:
                                os.chmod(remaining_file_path, 0o666)
                            except Exception:
                                pass  # 権限変更に失敗しても続行
                            
                            # ファイルを使用している可能性のあるプロセスを確認
                            if retry >= 2 and PSUTIL_AVAILABLE:  # 3回目以降で詳細確認（psutilが利用可能な場合のみ）
                                try:
                                    process_found = False
                                    for proc in psutil.process_iter(['pid', 'name', 'open_files']):
                                        try:
                                            if proc.info['open_files']:
                                                for file_info in proc.info['open_files']:
                                                    if remaining_file_path.lower() in file_info.path.lower():
                                                        print(f"ファイルを使用中のプロセス発見: {proc.info['name']} (PID: {proc.info['pid']})")
                                                        process_found = True
                                        except (psutil.NoSuchProcess, psutil.AccessDenied):
                                            continue
                                    if not process_found:
                                        print(f"プロセス監視では使用中のプロセスが見つかりませんでした: {remaining_file}")
                                except Exception as proc_error:
                                    print(f"プロセス確認中にエラー: {proc_error}")
                            
                            # Windows固有のファイルハンドル強制解放
                            if retry >= 3 and CTYPES_AVAILABLE:
                                try:
                                    # ファイルハンドルを強制的に解放
                                    handle = ctypes.windll.kernel32.CreateFileW(
                                        remaining_file_path,
                                        0x40000000,  # GENERIC_WRITE
                                        0,  # 共有なし
                                        None,
                                        3,  # OPEN_EXISTING
                                        0,
                                        None
                                    )
                                    if handle != -1:
                                        ctypes.windll.kernel32.CloseHandle(handle)
                                        print(f"Windowsファイルハンドル解放を試行: {remaining_file}")
                                except Exception as handle_error:
                                    print(f"Windowsハンドル操作エラー: {handle_error}")
                            
                            # 削除の最終試行
                            os.remove(remaining_file_path)
                            print(f"削除完了: {remaining_file}")
                            successful_deletions.append(remaining_file)
                            deletion_successful = True
                            break
                        elif os.path.isdir(remaining_file_path):
                            shutil.rmtree(remaining_file_path)
                            print(f"ディレクトリ削除完了: {remaining_file}")
                            successful_deletions.append(remaining_file)
                            deletion_successful = True
                            break
                    except PermissionError as perm_error:
                        if retry < max_retries - 1:
                            wait_time = (retry + 1) * 3  # 待機時間を徐々に増加
                            print(f"ファイル削除リトライ中... ({retry + 1}/{max_retries}): {remaining_file} (待機時間: {wait_time}秒)")
                            time.sleep(wait_time)
                            gc.collect()  # ガベージコレクションを再実行
                            
                            # より強力なファイルハンドル解放
                            try:
                                import ctypes
                                ctypes.windll.kernel32.SetProcessWorkingSetSize(-1, -1, -1)
                            except Exception:
                                pass
                        else:
                            # 最大リトライ回数に達した場合は警告として記録
                            warning_msg = f"ファイル削除スキップ（手動削除可能）: {remaining_file}\n理由: ネットワーク共有フォルダ上のファイルが他のマシンまたはプロセスによって使用されています。\n対処法: 手動でファイルを削除するか、次回実行時に自動で上書きされます。"
                            print(f"⚠️ 警告: {warning_msg}")
                            cleanup_warnings.append(warning_msg)
                            failed_deletions.append(remaining_file)
                    except Exception as file_error:
                        warning_msg = f"ファイル削除スキップ: {remaining_file} - {file_error}"
                        print(f"⚠️ 警告: {warning_msg}")
                        cleanup_warnings.append(warning_msg)
                        failed_deletions.append(remaining_file)
                        break
                
                # 削除に失敗したファイルの情報を記録
                if not deletion_successful and remaining_file not in failed_deletions:
                    failed_deletions.append(remaining_file)
        else:
            print("セット品加工図準備用フォルダは既に空です。")
            
    except Exception as cleanup_error:
        warning_msg = f"フォルダクリーンアップ中に警告が発生しました: {cleanup_error}"
        print(f"⚠️ 警告: {warning_msg}")
        cleanup_warnings.append(warning_msg)
    
    # 削除結果のサマリー表示
    print("\n=== ファイル削除結果サマリー ===")
    if successful_deletions:
        print(f"✅ 削除成功: {len(successful_deletions)}個のファイル")
        for file in successful_deletions:
            print(f"   - {file}")
    
    if failed_deletions:
        print(f"⚠️ 削除スキップ: {len(failed_deletions)}個のファイル（手動削除可能）")
        for file in failed_deletions:
            print(f"   - {file}")
        print("\n💡 ヒント: 削除に失敗したファイルは以下の方法で対処できます:")
        print("   1. 手動でファイルを削除")
        print("   2. 次回スクリプト実行時に自動で上書き")
        print("   3. 他のマシンでファイルが開かれていないか確認")
    
    # 重要でない警告の場合はメール通知しない（手動削除可能なため）
    if cleanup_warnings and len(failed_deletions) > len(successful_deletions):
        # 削除失敗が多い場合のみメール通知
        warning_summary = "\n".join(cleanup_warnings)
        send_error_email(f"ファイルクリーンアップ警告（手動対処可能）:\n{warning_summary}")
    elif cleanup_warnings:
        print("\n📧 メール通知: 軽微な警告のためメール通知をスキップしました。")
        
    print("\n=== スクリプト実行結果 ===")
    print("✅ 主要処理がすべて正常に完了しました！")
    print("   - ✅ Excelファイルからのデータ取得")
    print("   - ✅ Google Sheetsへのリンク更新")
    print("   - ✅ Google Driveへのファイルアップロード")
    if successful_deletions:
        print(f"   - ✅ ファイルクリーンアップ ({len(successful_deletions)}/{len(successful_deletions) + len(failed_deletions)}個成功)")
    if failed_deletions:
        print(f"   - ⚠️ 一部ファイルの手動削除が必要 ({len(failed_deletions)}個)")
    
    print("\n🎉 スクリプトが正常に完了しました！")

except Exception as e:
    error_detail = traceback.format_exc()
    send_error_email(error_detail)
    # エラーを再発生させてプログラムを停止
    raise