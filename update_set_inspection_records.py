import openpyxl
import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
import pyodbc as pyo
import smtplib
from email.mime.text import MIMEText
import traceback
import os
import sys
from dotenv import load_dotenv

# .envファイルから環境変数を読み込み
load_dotenv()


# 通知アカウント
# ・takada@araiseimitsu.onmicrosoft.com
# ・imai@araiseimitsu.onmicrosoft.com
# ・n.kizaki@araiseimitsu.onmicrosoft.com

# --- Google API設定 ---
GOOGLE_SERVICE_ACCOUNT_KEY_FILE = os.getenv("GOOGLE_SERVICE_ACCOUNT_KEY_FILE")

# --- メール通知用の設定 ---
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVERS = os.getenv("EMAIL_RECEIVERS", "").split(",") if os.getenv("EMAIL_RECEIVERS") else []
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.office365.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))

# Google API設定の存在確認
if not GOOGLE_SERVICE_ACCOUNT_KEY_FILE:
    raise ValueError("GOOGLE_SERVICE_ACCOUNT_KEY_FILE が .env ファイルに設定されていません")

# Google API認証ファイルパスの解決（環境非依存）
def resolve_google_api_key_file(filename):
    """
    Google API認証ファイルのパスを環境非依存で解決する
    """
    # 1. 現在の作業ディレクトリからの相対パス
    if os.path.exists(filename):
        return os.path.abspath(filename)
    
    # 2. スクリプトと同じディレクトリ
    script_dir = os.path.dirname(os.path.abspath(__file__))
    script_relative_path = os.path.join(script_dir, filename)
    if os.path.exists(script_relative_path):
        return script_relative_path
    
    # 3. ファイルが見つからない場合
    raise FileNotFoundError(
        f"Google API認証ファイルが見つかりません: {filename}\n"
        f"検索パス:\n"
        f"  - 現在の作業ディレクトリ: {os.path.abspath(filename)}\n"
        f"  - スクリプトディレクトリ: {script_relative_path}"
    )

# Google API認証ファイルパスを解決
RESOLVED_GOOGLE_API_KEY_FILE = resolve_google_api_key_file(GOOGLE_SERVICE_ACCOUNT_KEY_FILE)

# メール設定の存在確認
if not EMAIL_SENDER:
    raise ValueError("EMAIL_SENDER が .env ファイルに設定されていません")
if not EMAIL_PASSWORD:
    raise ValueError("EMAIL_PASSWORD が .env ファイルに設定されていません")
if not EMAIL_RECEIVERS:
    raise ValueError("EMAIL_RECEIVERS が .env ファイルに設定されていません")

def send_error_email(error_info):
    """
    エラー発生時に指定されたアカウントへメールを送信する関数
    """
    try:
        # プログラム名とファイルパスを取得
        program_name = os.path.basename(sys.argv[0])
        file_path = os.path.abspath(sys.argv[0])
        
        # 件名にプログラム名を追記
        subject = f"【エラー通知】{program_name} 実行中にエラーが発生しました"
        
        # 本文にプログラム名とファイルパスを追記
        body = f"""
お疲れ様です。

Pythonスクリプトの実行中にエラーが発生しました。
下記に詳細を記載します。

---
プログラム名: {program_name}

ファイルパス: {file_path}

日時: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')}

エラー詳細:
{error_info}
---

お手数ですが、ご確認をお願いします。
"""
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = EMAIL_SENDER
        msg["To"] = ", ".join(EMAIL_RECEIVERS)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECEIVERS, msg.as_string())
        print("エラー通知メールを送信しました。")

    except Exception as e:
        print(f"メール送信中にエラーが発生しました: {e}")

# --- メイン処理 ---
try:
    # セット済記録_Excel
    source_wb = openpyxl.load_workbook(r"\\192.168.1.200\共有\生産管理課\セット済記録.xlsx")
    source_ws = source_wb["セット記録原本"]

    # セット済記録_Excel_最終行取得
    last_row = source_ws.max_row
    for _ in range(0 + 1,last_row + 1):
        if source_ws.cell(_,1).value is not None:
            source_last_row = _

    # セット記録_データ取得
    data_col = [4, 13, 5, 8, 9, 10]
    set_data = []
    for row in range(source_last_row-10, source_last_row+1):
        row_data = []
        for col in data_col:
            value1 = source_ws.cell(row, col).value
            if isinstance(value1, datetime.datetime):
                value1 = datetime.datetime.strftime(value1, "%Y/%m/%d")
            row_data.append(value1)
        set_data.append(row_data)

    source_wb.close()

    # 今日の日付を取得
    today = datetime.datetime.today().strftime('%Y/%m/%d')
    # 今日の日付のみのリストを作成
    filtered_data = [row for row in set_data if row[0] == today]

    print(filtered_data)

    # 認証情報を設定（環境変数から読み込み）
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/spreadsheets",
             "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(RESOLVED_GOOGLE_API_KEY_FILE, scope)
    client = gspread.authorize(creds)

    # 品質管理記録
    ws = client.open_by_key("1se9Gvgd5dsk3hLAaLQ4C1M_yBsQpLBkZxWmHxEigrf0")  # スプレッドシートのkey
    sh_manage = ws.worksheet("セット品検査記録")
    sh_record = ws.worksheet("検査記録集計")

    # セット品検査記録シートデータ取得
    all_manage = sh_manage.get_all_values()

    # 完了したタスクを格納するリスト
    completed_tasks = []

    # リスト内の各行をチェック
    for task in all_manage[:]:  # コピーしたリストを使ってイテレーション
        if task[14] == "済":
            completed_tasks.append(task[0:14])  # [0]～[13] 番目のデータを追加
            all_manage.remove(task)

    # 検査記録集計シートデータ取得
    all_record = sh_record.get_all_values()

    # completed_tasks を all_record に結合
    all_record.extend(completed_tasks)

    # 検査記録集計シートデータ更新
    sh_record.update(values=all_record, range_name="A1")

    # all_manage を分割
    data_0_to_10 = [row[:11] for row in all_manage]
    data_12_to_13 = [row[12:14] for row in all_manage]

    # set_data を data_0_to_10 に結合
    data_0_to_10.extend(filtered_data)

    # データのフィルタリング
    filter_data_0_to_10 = [row for row in data_0_to_10 if row[0] != '']
    filter_data_0_to_10.insert(0, data_0_to_10[0])  # ヘッダーを追加

    # 最初の行を削除
    del filter_data_0_to_10[0]

    # セル範囲A2:K150, M2:O150を空白にする
    sh_manage.update(range_name="A2:K150", values=[[""] * 11] * 149)
    sh_manage.update(range_name="M2:O150", values=[[""] * 3] * 149)

    # A1セルにdata_0_to_10のデータを書き込む
    sh_manage.update(range_name="A1", values=filter_data_0_to_10)

    # M1セルにdata_12_to_13のデータを書き込む
    sh_manage.update(range_name="M1", values=data_12_to_13)

    print("complete")
    
except Exception as e:
    error_detail = traceback.format_exc()
    send_error_email(error_detail)
    # エラーを再発生させてプログラムを停止
    raise