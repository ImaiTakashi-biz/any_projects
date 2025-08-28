import openpyxl
import datetime
import gspread
import json
from oauth2client.service_account import ServiceAccountCredentials
import os
import smtplib
from email.mime.text import MIMEText
import traceback
import sys
from dotenv import load_dotenv

import pyodbc as pyo
import tkinter as tk
from tkinter import ttk

# .envファイルから環境変数を読み込み
load_dotenv()


# 通知アカウント
# ・takada@araiseimitsu.onmicrosoft.com
# ・imai@araiseimitsu.onmicrosoft.com
# ・n.kizaki@araiseimitsu.onmicrosoft.com

# --- Google API設定 ---
GOOGLE_SERVICE_ACCOUNT_KEY_FILE = os.getenv("GOOGLE_SERVICE_ACCOUNT_KEY_FILE")

# --- メール通知用の設定 ---
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVERS = os.getenv("EMAIL_RECEIVERS", "").split(",") if os.getenv("EMAIL_RECEIVERS") else []
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.office365.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))

# Google API設定の存在確認
if not GOOGLE_SERVICE_ACCOUNT_KEY_FILE:
    raise ValueError("GOOGLE_SERVICE_ACCOUNT_KEY_FILE が .env ファイルに設定されていません")

# Google API認証ファイルパスの解決（環境非依存）
def resolve_google_api_key_file(filename):
    """
    Google API認証ファイルのパスを環境非依存で解決する
    """
    # 1. 現在の作業ディレクトリからの相対パス
    if os.path.exists(filename):
        return os.path.abspath(filename)
    
    # 2. スクリプトと同じディレクトリ
    script_dir = os.path.dirname(os.path.abspath(__file__))
    script_relative_path = os.path.join(script_dir, filename)
    if os.path.exists(script_relative_path):
        return script_relative_path
    
    # 3. ファイルが見つからない場合
    raise FileNotFoundError(
        f"Google API認証ファイルが見つかりません: {filename}\n"
        f"検索パス:\n"
        f"  - 現在の作業ディレクトリ: {os.path.abspath(filename)}\n"
        f"  - スクリプトディレクトリ: {script_relative_path}"
    )

# Google API認証ファイルパスを解決
RESOLVED_GOOGLE_API_KEY_FILE = resolve_google_api_key_file(GOOGLE_SERVICE_ACCOUNT_KEY_FILE)

# メール設定の存在確認
if not EMAIL_SENDER:
    raise ValueError("EMAIL_SENDER が .env ファイルに設定されていません")
if not EMAIL_PASSWORD:
    raise ValueError("EMAIL_PASSWORD が .env ファイルに設定されていません")
if not EMAIL_RECEIVERS:
    raise ValueError("EMAIL_RECEIVERS が .env ファイルに設定されていません")

def send_error_email(error_info):
    """
    エラー発生時に指定されたアカウントへメールを送信する関数
    """
    try:
        # プログラム名とファイルパスを取得
        program_name = os.path.basename(sys.argv[0])
        file_path = os.path.abspath(sys.argv[0])
        
        # 件名にプログラム名を追記
        subject = f"【エラー通知】{program_name} 実行中にエラーが発生しました"
        
        # 本文にプログラム名とファイルパスを追記
        body = f"""
お疲れ様です。

Pythonスクリプトの実行中にエラーが発生しました。
下記に詳細を記載します。

---
プログラム名: {program_name}

ファイルパス: {file_path}

日時: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')}

エラー詳細:
{error_info}
---

お手数ですが、ご確認をお願いします。
"""
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = EMAIL_SENDER
        msg["To"] = ", ".join(EMAIL_RECEIVERS)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECEIVERS, msg.as_string())
        print("エラー通知メールを送信しました。")

    except Exception as e:
        print(f"メール送信中にエラーが発生しました: {e}")


def update_p_bar(current_value, max_value):
    pb["value"] = current_value / max_value * 100
    pb_label.config(text=f"進捗率 :{pb['value']:.0f} %")
    pb_label.pack(pady=20)
    pb.pack(pady=20)
    window.update()
    if current_value >= max_value:
        window.destroy()
    else:
        pass


# --- メイン処理 ---
try:
    window = tk.Tk()
    window.title("セット予定品不具合情報更新中.....")
    pb = ttk.Progressbar(window, maximum= 100, length= 500, mode="determinate")
    pb_label = tk.Label(window, text="進捗状況 : 0%", font=50)
    pb_label.pack(pady=20)
    pb.pack(pady=20)

    update_p_bar(current_value=1, max_value=13)
    wb1 = openpyxl.load_workbook(r"\\192.168.1.200\共有\製造課\ロボパット\ロボパット用.xlsx",data_only=True)
    ws1 = wb1["Data"]
    wb2 = openpyxl.load_workbook(r"\\192.168.1.200\共有\製造課\ロボパット\セット品不具合情報.xlsx")
    ws2 = wb2.worksheets[0]

    last_row = ws1.max_row
    print(last_row)

    r = 2
    No_list = []
    for i in range(36, 45):
        if ws1.cell(i, 7).value == None:
            continue
        No_list.append(ws1.cell(12, r).value)
        r += 1
    wb1.close()
    print(No_list)

    update_p_bar(current_value=2, max_value=13)
    #print(pyo.drivers())
    #access_db接続
    for driver in pyo.drivers():
        if driver.startswith('Microsoft Access Driver'):
            print(driver)
    con_str = (
        r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=\\192.168.1.200\共有\品質保証課\外観検査記録\不具合情報検索.accdb'
        )
    con = pyo.connect(con_str)
    cursor = con.cursor()
    for table in cursor.tables(tableType='TABLE'):
        print(table.table_name)

    # for row in ws2.iter_rows(min_row=2):
    #   for cell in row:
    #   	 cell.value = None

    update_p_bar(current_value=3, max_value=13)
    No_list_len = len(No_list)
    row = 2
    for r, i in enumerate(No_list):
        sql = "select * FROM t_不具合情報 WHERE 品番= '" + str(i) + "' order by 生産ロットID"
        cursor.execute(sql)
        data = cursor.fetchall()
        for l in data:
            print(type(l.指示日))
            print(l.生産ロットID,l.指示日,l.品番,l.号機,l.数量,l.総不具合数,l.不良率,l.外観キズ,l.圧痕,l.切粉,l.毟れ,l.穴大,l.穴小,l.穴キズ,l.バリ,l.短寸,l.面粗,l.サビ,l.ボケ,l.挽目,l.汚れ,l.メッキ,l.落下,l.フクレ,l.ツブレ,l.ボッチ,l.段差,l.バレル石,l.径プラス,l.径マイナス,l.ゲージ,l.異物混入,l.形状不良,l.こすれ,l.変色シミ,l.材料キズ,l.ゴミ,l.その他,l.その他内容,l.検査者1,l.検査者2,l.検査者3,l.検査者4,l.検査者5,l.時間)
            ws2.cell(row, 1, value=l.生産ロットID)
            # day = datetime.datetime.date(l. 指示日)
            ws2.cell(row, 2, value=l.指示日)
            ws2.cell(row, 3, value=str(l.品番))
            ws2.cell(row, 4, value =l.号機)
            ws2.cell(row, 5, value = l.数量)
            ws2.cell(row, 6, value = int(l.総不具合数))
            ws2.cell(row, 7, value = l.不良率)
            ws2.cell(row, 8, value = l.外観キズ)
            ws2.cell(row, 9, value = l.圧痕)
            ws2.cell(row, 10, value = l.切粉)
            ws2.cell(row, 11, value = l.毟れ)
            ws2.cell(row, 12, value = l.穴大)
            ws2.cell(row, 13, value = l.穴小)
            ws2.cell(row, 14, value = l.穴キズ)
            ws2.cell(row, 15, value = l.バリ)
            ws2.cell(row, 16, value = l.短寸)
            ws2.cell(row, 17, value = l.面粗)
            ws2.cell(row, 18, value = l.サビ)
            ws2.cell(row, 19, value = l.ボケ)
            ws2.cell(row, 20, value = l.挽目)
            ws2.cell(row, 21, value = l.汚れ)
            ws2.cell(row, 22, value = l.メッキ)
            ws2.cell(row, 23, value = l.落下)
            ws2.cell(row, 24, value = l.フクレ)
            ws2.cell(row, 25, value = l.ツブレ)
            ws2.cell(row, 26, value = l.ボッチ)
            ws2.cell(row, 27, value = l.段差)
            ws2.cell(row, 28, value = l.バレル石)
            ws2.cell(row, 29, value = l.径プラス)
            ws2.cell(row, 30, value = l.径マイナス)
            ws2.cell(row, 31, value = l.ゲージ)
            ws2.cell(row, 32, value = l.異物混入)
            ws2.cell(row, 33, value = l.形状不良)
            ws2.cell(row, 34, value = l.こすれ)
            ws2.cell(row, 35, value = l.変色シミ)
            ws2.cell(row, 36, value = l.材料キズ)
            ws2.cell(row, 37, value = l.ゴミ)
            ws2.cell(row, 38, value = l.その他)
            ws2.cell(row, 39, value = l.その他内容)
            ws2.cell(row, 40, value = l.検査者1)
            ws2.cell(row, 41, value = l.検査者2)
            ws2.cell(row, 42, value = l.検査者3)
            ws2.cell(row, 43, value = l.検査者4)
            ws2.cell(row, 44, value = l.検査者5)
            ws2.cell(row, 45, value = l.時間)
            row += 1

    wb2.save(r"\\192.168.1.200\共有\製造課\ロボパット\セット品不具合情報"
             r".xlsx")
    wb2.close()
    update_p_bar(current_value=4, max_value=13)


    # 認証情報を設定（環境変数から読み込み）
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/spreadsheets",
             "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(RESOLVED_GOOGLE_API_KEY_FILE, scope)
    client = gspread.authorize(creds)

    # スプレッドシートを開く
    # ws = client.open("name").sheet1 #スプレッドシートの名前
    ws = client.open_by_key("1csorahdNKva2papZSkW8-BtRPX8C6cvLrulQxhym5Ik") #スプレッドシートのkey
    sh = ws.worksheet("工程内トラブル品報告")

    # 工程内トラブル品報告　データ取得
    all_values = sh.get_all_values()
    update_p_bar(current_value=5, max_value=13)

    # ロボパット用.xlsx
    wb_rob = openpyxl.load_workbook(r"\\192.168.1.200\共有\製造課\ロボパット\ロボパット用.xlsx", read_only=True)
    ws_rob = wb_rob["Data"]

    # ロボパット用データ取得
    rob_data = []
    for col in range(2, 12):
        row_list = []
        for row in range(8, 32):
            value1 = ws_rob.cell(row, col).value
            if isinstance(value1, datetime.datetime):
                value1 = datetime.datetime.strftime(value1, "%Y/%m/%d")
            row_list.append(value1)
        rob_data.append(row_list)
    # #品番のみ
    # rob_data = []
    # for col in range(2, 12):
    #   value1 = ws_rob.cell(12, col).value
    #   rob_data.append(value1)
    wb_rob.close()
    update_p_bar(current_value=6, max_value=13)

    # セット品不具合情報.xlsx
    wb_defect = openpyxl.load_workbook(r"\\192.168.1.200\共有\製造課\ロボパット\セット品不具合情報.xlsx")
    ws_defect = wb_defect["Sheet1"]

    # セット品不具合情報データ取得
    last_row = ws_defect.max_row
    for _ in range(0 + 1, last_row + 1):
        if ws_defect.cell(_, 1).value is not None:
            last_row1 = _

    defect_data = []
    for row in range(2, last_row1 + 1):
        row_list1 = []
        for col in range(1, 46):
            value2 = ws_defect.cell(row, col).value
            if isinstance(value2, datetime.datetime):
                value2 = datetime.datetime.strftime(value2, "%Y/%m/%d")
            row_list1.append(value2)
        defect_data.append(row_list1)
    wb_defect.close()
    update_p_bar(current_value=7, max_value=13)

    # セット予定表.xlsx
    wb_set = openpyxl.load_workbook(r"\\192.168.1.200\共有\生産管理課\セット予定表.xlsx",read_only=True, data_only=True)
    ws_set = wb_set["生産中"]

    # 生産中データ取得
    set_data = []
    target_col = [5, 6, 8, 9, 10, 11, 12, 15, 57, 19, 21, 48, 56]
    for row in range(2, 72):
        data = []
        for col in target_col:
            value_data = ws_set.cell(row, col).value
            if isinstance(value_data, datetime.datetime):
                value_data = datetime.datetime.strftime(value_data, "%Y/%m/%d")
            data.append(value_data)
        set_data.append(data)
    wb_set.close()
    update_p_bar(current_value=8, max_value=13)

    # セット予定品不具合情報更新用.xlsx
    wb = openpyxl.load_workbook(r"\\192.168.1.200\共有\製造課\ロボパット\セット予定品不具合情報更新用.xlsx")
    ws = wb["セット品"]
    ws1 = wb["加工中リンク"]

    # 加工中リンク更新
    target_col1 = [1, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14]
    row = 2
    for set in set_data:
        a = 0
        for col in target_col1:
            ws1.cell(row, col).value = set[a]
            a += 1
        row += 1
    update_p_bar(current_value=9, max_value=13)

    # セットデータ更新
    target_row = [1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25]
    col = 2
    for rob in rob_data:
        b = 0
        for row in target_row:
            ws.cell(row, col).value = rob[b]
            b += 1
        col += 1
    update_p_bar(current_value=10, max_value=13)

    # 外観検査情報書き込み
    r = 1
    for product in rob_data:
        sheet_name = str(r)
        ws_write = wb[sheet_name]
        # データクリア
        clear_range = ws_write['A3:AS200']
        for row in clear_range:
            for cell in row:
                cell.value = None
        # データ書き込み
        r1 = 3
        for defect in defect_data:
            if product[4] == defect[2]:
                for col in range(1, 46):
                    ws_write.cell(r1, col).value = defect[col - 1]
                r1 += 1
        r += 1
    update_p_bar(current_value=11, max_value=13)

    # 工程内トラブル情報書き込み
    r = 11
    for product in rob_data:
        sheet_name = str(r)
        ws_write = wb[sheet_name]
        # データクリア
        clear_range = ws_write['A3:AS200']
        for row in clear_range:
            for cell in row:
                cell.value = None
        # データ書き込み
        r1 = 3
        for trouble in all_values:
            if product[4] == trouble[8]:
                for col in range(1, 30):
                    ws_write.cell(r1, col).value = trouble[col - 1]
                r1 += 1
        r += 1
    update_p_bar(current_value=12, max_value=13)

    # 変化点・品質情報　ファイル追加
    for _ in rob_data:
        if _[0] is not None:
            if os.path.exists(rf"\\192.168.1.200\共有\品質保証課\変化点・品質情報\{_[4]}.xlsx"):
                pass
            else:
                wb_new = openpyxl.load_workbook(r"\\192.168.1.200\共有\品質保証課\変化点・品質情報\1.変化点・品質情報-原紙.xlsx")
                ws_new = wb_new["main"]
                ws_new.cell(3, 1).value = _[20]
                ws_new.cell(3, 4).value = _[4]
                ws_new.cell(3, 8).value = _[5]
                new = rf"\\192.168.1.200\共有\品質保証課\変化点・品質情報\{_[4]}.xlsx"
                wb_new.save(new)
                wb_new.close()
        else:
            pass

    wb.save(r"\\192.168.1.200\共有\製造課\ロボパット\セット予定品不具合情報更新用.xlsx")
    wb.save(r"\\192.168.1.200\共有\製造課\工程内トラブル品報告\セット予定品不具合情報.xlsx")
    wb.close()
    update_p_bar(current_value=13, max_value=13)
    
    print("完了しました。")

except Exception as e:
    error_detail = traceback.format_exc()
    send_error_email(error_detail)
    # エラーを再発生させてプログラムを停止
    raise